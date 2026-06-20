#order_views.py
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Count
from django.db import transaction
from api.services.qr_service import bulk_generate_qr_codes
from api.services.id_engine import finalize_order_production
from api.services.order_service import create_full_order
from rest_framework.decorators import api_view, permission_classes, parser_classes
from api.models.orders import Order
from api.models.students import Student
from api.services.processing_service import start_processing_queue
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.conf import settings
from api.models.user_profile import UserProfile
from api.utils.permissions import check_order_access, check_student_access
import os
import zipfile
from io import BytesIO
import threading
import json
import requests
import cloudinary


# Download all QR codes for an order as a ZIP file
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_qr_codes(request, order_id):
    order, err = check_order_access(request, order_id)
    if err:
        return err

    try:
        students = Student.objects.filter(order=order).exclude(qr_code_url__isnull=True).exclude(qr_code_url='')
        
        if not students.exists():
            return Response({'error': 'No QR codes found'}, status=404)
        
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for student in students:
                if student.qr_code_url:
                    try:
                        response = requests.get(student.qr_code_url, timeout=10)
                        if response.status_code == 200 and len(response.content) > 0:
                            clean_name = student.full_name.replace(' ', '_')
                            filename = f"{student.student_id}_{clean_name}.png"
                            zip_file.writestr(filename, response.content)
                        else:
                            print(f"Warning: Could not download QR for student {student.id}")
                    except Exception as e:
                        print(f"Error downloading QR for student {student.id}: {e}")
        
        zip_buffer.seek(0)
        content = zip_buffer.getvalue()
        
        if not content or len(content) < 100:
            return Response({'error': 'Generated ZIP was empty'}, status=500)
        
        response = HttpResponse(content, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="qr_codes_order_{order_id}.zip"'
        return response

    except Exception as e:
        return Response({'error': str(e)}, status=500)

# Get single student QR code
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_qr(request, student_id):
    student, err = check_student_access(request, student_id)
    if err:
        return err
    if not student.qr_code_image:
        return Response({'error': 'QR code not generated yet'}, status=404)
    return Response(student.qr_code_image.read(), content_type='image/png')

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def order_controller(request):
    if request.method == 'GET':
        return _list_orders(request)
    return _create_order(request)

def _list_orders(request):
    try:
        try:
            profile = request.user.profile
        except UserProfile.DoesNotExist:
            return Response({'error': 'User profile not found.'}, status=403)

        # Default to nothing — never leak data
        orders = Order.objects.none()

        if profile.role == UserProfile.Role.ADMIN:
            orders = Order.objects.annotate(total_students=Count('students'))

        elif profile.role == UserProfile.Role.OPERATOR:
            orders = Order.objects.annotate(
                total_students=Count('students')
            ).filter(assigned_operator=request.user)

        elif profile.role in (UserProfile.Role.INSTITUTION, UserProfile.Role.COORDINATOR):
            if not profile.institution:
                return Response({'error': 'No institution linked to this account.'}, status=403)
            orders = Order.objects.annotate(
                total_students=Count('students')
            ).filter(institution=profile.institution)

        result = orders.values(
            'id', 'school_name', 'batch_name', 'status', 'created_at', 'total_students'
        )

        order_list = [
            {**{k: v for k, v in o.items() if k != 'total_students'}, 'student_count': o['total_students']}
            for o in result
        ]
        return Response(order_list)

    except Exception as e:
        return Response({'error': str(e)}, status=500)
    
def _create_order(request):
    try:
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return Response({'error': 'Invalid JSON format'}, status=400)

        school_name = data.get('school_name')
        batch_name = data.get('batch_name')
        students = data.get('students', [])

        if not school_name or not batch_name:
            return Response({'error': 'Missing school/batch name'}, status=400)

        if not isinstance(students, list) or len(students) == 0:
            return Response({'error': 'Student list is empty'}, status=400)
        
        try:
            institution = request.user.institution
        except Exception:
            institution = None

        new_order = create_full_order(
            school_name=school_name,
            batch_name=batch_name,
            student_list=students,
            institution=institution
        )

        # Return the same structure as _list_orders so the UI doesn't break
        return JsonResponse({
            'id': new_order.id, #type: ignore
            'school_name': new_order.school_name,
            'batch_name': new_order.batch_name,
            'status': new_order.status,
            'student_count': len(students),
            'qr_generated': new_order.student_count,
            'message': 'Order created successfully.'
        }, status=201)

    except Exception as e:
        return Response({'error': str(e)}, status=400)
    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def start_processing(request, pk):
    order, err = check_order_access(request, pk)
    if err:
        return err
    try:
        updated = Order.objects.filter(id=pk, status__in=['PENDING', 'FAILED', 'PROOFING']).update(status='PROCESSING')
        if not updated:
            return Response({'error': 'Order is already processing or does not exist'}, status=400)

        thread = threading.Thread(target=start_processing_queue, args=(pk,))
        thread.daemon = True
        thread.start()
        return Response({'message': 'Processing started in the background', 'order_id': pk})
    except Exception as e:
        return Response({'error': str(e)}, status=500)
    
@api_view(['POST']) 
@permission_classes([IsAuthenticated])
def complete_order(request, pk):
    order, err = check_order_access(request, pk)
    if err:
        return err
    try:
        success, new_status = finalize_order_production(pk)
        if success:
            if new_status == 'COMPLETED':
                Order.objects.filter(id=pk).update(completed_at=timezone.now())
            return JsonResponse({'message': f'Order moved to {new_status}'}, status=200)
        return JsonResponse({"error": "Order is not in a state that can be completed"}, status=400)
    except Exception as e:
        return Response({"error": str(e)}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_photos(request, order_id, student_id=None):
    order, err = check_order_access(request, order_id)
    if err:
        return err
    
    if student_id is not None:

        from api.services.id_engine import manual_crop_photo as execute_manual_crop
        return execute_manual_crop(request, order_id, student_id)
        
    if order.status not in [Order.Status.PENDING, Order.Status.FAILED, Order.Status.PROOFING]:
        return Response({ 'error': 'Photos can only be uploaded when order is PENDING, PROOFING or FAILED'}, status=400)

    files = request.FILES.getlist('files')

    if not files:
        return Response({'error': 'No files provided'}, status=400)
    
    uploaded = []
    failed = []

    for file in files: 
        try: 
            upload_result = cloudinary.uploader.upload(
                file,
                folder=f'student_photos/order_{order_id}',
                resource_type='image',
                use_filename=True,
                unique_filename=False
            )
            uploaded.append({
                'file': file.name,
                'url': upload_result['secure_url'],
                'public_id': upload_result['public_id']
            })
        except Exception as e:
            failed.append({'file': file.name, 'error': str(e)})
        
    return Response({
        'message': f'{len(uploaded)} photo(s) uploaded successfully',
        'uploaded': uploaded,
        'failed': failed
    }, status=207 if failed else 201)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def parse_order_file(request):
    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'No file provided'}, status=400)

    file_name = file.name.lower()
    sheet_name = request.data.get('sheet_name', None)

    COLUMN_MAP = {
        'name':       ['name', 'full name', 'full_name'],
        'student_id': ['student_id', 'student id', 'studentid'],
        'grade':      ['grade', 'grade level', 'grade_level'],
        'section':    ['section'],
    }

    def map_row(row: dict) -> dict:
        result = {}
        row_lower = {k.strip().lower(): v for k, v in row.items()}
        for standard_key, variants in COLUMN_MAP.items():
            for variant in variants:
                if variant in row_lower:
                    result[standard_key] = str(row_lower[variant]).strip() if row_lower[variant] else ''
                    break
            else:
                result[standard_key] = ''
        return result

    try:
        # ── CSV ───────────────────────────────────────────────────────────────
        if file_name.endswith('.csv'):
            import csv, io
            decoded = file.read().decode('utf-8-sig')
            reader  = csv.DictReader(io.StringIO(decoded))

            # ✅ Strip BOM and whitespace from headers
            reader.fieldnames = [
                f.strip().lstrip('\ufeff') for f in (reader.fieldnames or [])
            ]

            rows = [map_row(row) for row in reader]

            return Response({
                'rows':       rows,
                'sheets':     [],
                'total_rows': len(rows),
                'file_name':  file.name,
            })

        # ── xlsx ──────────────────────────────────────────────────────────────
        elif file_name.endswith('.xlsx'):
            from openpyxl import load_workbook

            file.seek(0)
            file_bytes = BytesIO(file.read())
            file_bytes.seek(0)  # ✅ seek BytesIO after writing

            wb          = load_workbook(file_bytes, read_only=True, data_only=True)
            sheet_names = wb.sheetnames

            if len(sheet_names) > 1 and not sheet_name:
                wb.close()
                return Response({
                    'rows':       [],
                    'sheets':     sheet_names,
                    'total_rows': 0,
                    'file_name':  file.name,
                })

            target_sheet = sheet_name if sheet_name in sheet_names else sheet_names[0]
            ws           = wb[target_sheet]
            rows_iter    = ws.iter_rows(values_only=True)
            headers      = [str(cell).strip() if cell else '' for cell in next(rows_iter)]

            rows = []
            for row in rows_iter:
                raw = {
                    headers[i]: str(row[i]).strip() if row[i] is not None else ''
                    for i in range(len(headers))
                }
                if not any(raw.values()):
                    continue
                rows.append(map_row(raw))

            wb.close()

            return Response({
                'rows':       rows,
                'sheets':     sheet_names,
                'total_rows': len(rows),
                'file_name':  file.name,
            })

        # ── xls ───────────────────────────────────────────────────────────────
        elif file_name.endswith('.xls'):
            import xlrd

            file.seek(0)
            wb          = xlrd.open_workbook(file_contents=file.read())
            sheet_names = wb.sheet_names()

            if len(sheet_names) > 1 and not sheet_name:
                return Response({
                    'rows':       [],
                    'sheets':     sheet_names,
                    'total_rows': 0,
                    'file_name':  file.name,
                })

            target_sheet = sheet_name if sheet_name in sheet_names else sheet_names[0]
            ws           = wb.sheet_by_name(target_sheet)

            if ws.nrows == 0:
                return Response({'error': 'The selected sheet is empty.'}, status=400)

            headers = [str(ws.cell_value(0, col)).strip() for col in range(ws.ncols)]

            rows = []
            for row_idx in range(1, ws.nrows):
                raw = {
                    headers[col]: str(ws.cell_value(row_idx, col)).strip()
                    for col in range(ws.ncols)
                }
                if not any(raw.values()):
                    continue
                rows.append(map_row(raw))

            return Response({
                'rows':       rows,
                'sheets':     sheet_names,
                'total_rows': len(rows),
                'file_name':  file.name,
            })

        else:
            return Response({'error': 'Unsupported file type.'}, status=400)

    except Exception as e:
        return Response({'error': f'Failed to parse file: {str(e)}'}, status=400)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def check_duplicate_order(request):
    """
    Checks if an order with the same school_name + batch_name already exists
    for this institution.

    Body:
        - school_name: str
        - batch_name: str
        - student_count: int

    Returns:
        - is_duplicate: bool
        - existing_order_id: int | null
        - existing_count: int | null
        - can_override: bool (true if new file has more rows than existing)
    """
    school_name = request.data.get('school_name', '').strip()
    batch_name = request.data.get('batch_name', '').strip()
    student_count = request.data.get('student_count', 0)
    
    if not school_name or not batch_name:
        return Response({'error': 'school_name and batch_name are required'}, status=400)
    
    try: 
        student_count = int(student_count)
    except (ValueError, TypeError):
        return Response({'error': 'student_count must be an integer'}, status=400)

    try:
        institution = request.user.institution
    except Exception:
        institution = None
    
    existing = Order.objects.filter(
        school_name__iexact=school_name,
        batch_name__iexact=batch_name,
        institution=institution
    ).exclude(
        status__in=[Order.Status.CANCELLED, Order.Status.FAILED]
    ).first()

    if not existing:
        return Response({
            'is_duplicate': False,
            'existing_order_id': None,
            'existing_count': None,
            'can_override': False,
        })
    
    can_override = student_count > existing.student_count

    return Response({
        'is_duplicate':      True,
        'existing_order_id': existing.id,
        'existing_count':    existing.student_count,
        'can_override':      can_override,
    })

@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_order(request, order_id):
    """
    Updates an existing order in place with a new student list.
    Used when a duplicate order is detected but the new file
    has more rows (override confirmed by user).

    Body:
        - school_name: str (optional)
        - batch_name: str (optional)
        - students: list of { name, student_id, grade, section }
    """
    order, err = check_order_access(request, order_id)
    if err:
        return err

    # Guard — only allow override on non-terminal statuses
    locked_statuses = [
        Order.Status.PRINTING,
        Order.Status.COMPLETED,
        Order.Status.CANCELLED,
    ]
    if order.status in locked_statuses:
        return Response({
            'error': f'Order cannot be updated in {order.status} status'
        }, status=400)

    students = request.data.get('students', [])
    if not isinstance(students, list) or len(students) == 0:
        return Response({'error': 'Student list is empty'}, status=400)

    # Optional field updates
    if 'school_name' in request.data:
        order.school_name = request.data['school_name'].strip()
    if 'batch_name' in request.data:
        order.batch_name = request.data['batch_name'].strip()

    try:
        with transaction.atomic():
            # Wipe existing students and re-create
            Student.objects.filter(order=order).delete()

            student_ids = [s['student_id'] for s in students]
            if len(student_ids) != len(set(student_ids)):
                raise ValueError('Duplicate student_id values in the new student list')

            student_objs = [
                Student(
                    order=order,
                    full_name=s['name'],
                    student_id=s['student_id'],
                    grade_level=s['grade'],
                    section=s.get('section', ''),
                ) for s in students
            ]
            Student.objects.bulk_create(student_objs)

            # Reset order back to PENDING with new count
            order.student_count = len(students)
            order.status = Order.Status.PENDING
            order.approved_by = None
            order.approved_at = None
            order.approval_notes = ''
            order.save()

            # Re-generate QR codes for new students
            new_students = Student.objects.filter(order=order)
            qr_result = bulk_generate_qr_codes(new_students, order.id)
            print(f"QR Re-generation: {qr_result['generated']} success, {qr_result['failed']} failed")

    except ValueError as e:
        return Response({'error': str(e)}, status=400)
    except Exception as e:
        return Response({'error': f'Update failed: {str(e)}'}, status=500)

    return Response({
        'id':            order.id,
        'school_name':   order.school_name,
        'batch_name':    order.batch_name,
        'status':        order.status,
        'student_count': order.student_count,
        'message':       'Order updated successfully.',
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_id_cards(request, order_id):
    order, err = check_order_access(request, order_id)
    if err:
        return err

    students = Student.objects.filter(
        order=order,
        photo_status=Student.PhotoStatus.PROCESSED
    ).exclude(processed_photo='').exclude(processed_photo__isnull=True)

    if not students.exists():
        return Response({'error': 'No processed ID cards found for this order'}, status=404)

    zip_buffer = BytesIO()
    added = 0

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for student in students:
            try:
                file_path = student.processed_photo.path
                if os.path.exists(file_path):
                    clean_name = student.full_name.replace(' ', '_')
                    ext = os.path.splitext(file_path)[1] or '.png'
                    filename = f"{student.student_id}_{clean_name}{ext}"
                    zip_file.write(file_path, filename)
                    added += 1
            except Exception as e:
                print(f"Skipping student {student.student_id}: {e}")

    if added == 0:
        return Response({'error': 'No ID card files found on disk'}, status=404)

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="id_cards_order_{order_id}.zip"'
    return response